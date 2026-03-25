import json
import os
from openpyxl import load_workbook
from typing import Dict, Any, List
from common.log_handler import logger

class DataReader:
    """测试数据读取工具类，支持JSON/Excel格式"""

    @staticmethod
    def read_json(file_name: str) -> Dict[str, Any]:
        """
        读取JSON格式测试数据
        :param file_name: 数据文件名（如"user_data.json"，默认在test_data目录）
        :return: 解析后的JSON数据
        """
        data_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "test_data", file_name)
        try:
            with open(data_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            logger.info(f"成功读取JSON测试数据：{data_path}")
            return data
        except FileNotFoundError:
            logger.error(f"JSON测试数据文件不存在：{data_path}")
            raise
        except json.JSONDecodeError as e:
            logger.error(f"JSON数据格式错误：{data_path}，异常信息：{str(e)}")
            raise

    @staticmethod
    def read_excel(file_name: str, sheet_name: str = None) -> List[List[Any]]:
        """
        读取Excel格式测试数据（返回二维列表，第一行为表头）
        :param file_name: 数据文件名（如"order_data.xlsx"，默认在test_data目录）
        :param sheet_name: 工作表名称（默认读取第一个工作表）
        :return: 解析后的Excel数据
        """
        data_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "test_data", file_name)
        try:
            workbook = load_workbook(data_path, data_only=True)
            # 选择工作表
            if sheet_name is None:
                sheet = workbook.active
            else:
                sheet = workbook[sheet_name]
            # 读取所有数据
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            workbook.close()
            logger.info(f"成功读取Excel测试数据：{data_path}，工作表：{sheet.title}，数据行数：{len(data)}")
            return data
        except FileNotFoundError:
            logger.error(f"Excel测试数据文件不存在：{data_path}")
            raise
        except Exception as e:
            logger.error(f"Excel数据读取失败：{data_path}，异常信息：{str(e)}")
            raise
